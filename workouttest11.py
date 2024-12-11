import json
from datetime import datetime, date, timedelta
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkthemes import ThemedTk
from tkcalendar import DateEntry
from PIL import Image, ImageTk, ImageDraw
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


def parse_workout_data(data):
    """
    Parse workout data from a text string and return a structured list.
    """
    workouts = []
    current_workout = None
    current_exercise = None
    for line in data.split('\n'):
        line = line.strip()
        if line.startswith('Day'):
            if current_workout:
                workouts.append(current_workout)
            current_workout = {'day': line, 'exercises': []}
            current_exercise = None
        elif line and current_workout:
            if 'x' in line and len(line.split('x')) == 2:
                sets, reps = line.split('x')
                if current_exercise:
                    current_workout['exercises'].append({
                        'exercise': current_exercise,
                        'sets': sets.strip(),
                        'reps': reps.strip(),
                        'weight': ''
                    })
                    current_exercise = None
                else:
                    parts = line.rsplit(' ', 2)
                    if len(parts) == 3:
                        exercise, sets, reps = parts
                        current_workout['exercises'].append({
                            'exercise': exercise.strip(),
                            'sets': sets.strip(),
                            'reps': reps.strip(),
                            'weight': ''
                        })
                    else:
                        current_exercise = line
            else:
                if current_exercise:
                    current_workout['exercises'].append({
                        'exercise': current_exercise,
                        'sets': '',
                        'reps': '',
                        'weight': ''
                    })
                current_exercise = line
    if current_workout:
        if current_exercise:
            current_workout['exercises'].append({
                'exercise': current_exercise,
                'sets': '',
                'reps': '',
                'weight': ''
            })
        workouts.append(current_workout)
    return workouts


class WorkoutPlanner:
    """
    A GUI application for planning and tracking workouts.
    """

    def __init__(self, root, workouts):
        """Initialize the WorkoutPlanner application."""
        self.root = root  # Use the root window passed in
        # Configure styles, etc.
        self.root.title("Workout Plan")
        self.root.geometry("1400x900")

        # Initialize styles
        self.style = ttk.Style(self.root)
        self.style.theme_use('arc')  # Set the theme

        # Data initialization
        self.workouts = self.restructure_workouts(workouts)
        self.total_weeks = len(self.workouts)
        self.current_week = 1
        self.current_day = 0

        self.prs = self.initialize_prs()

        self.muscle_engagement = self.initialize_muscle_engagement()
        self.weekly_muscle_engagement = self.initialize_weekly_muscle_engagement()
        self.start_date = self.find_start_date()
        self.current_date = date.today()

        self.recovery_rates = {
            'chest': 15, 'shoulders': 15, 'biceps': 20, 'triceps': 20,
            'abs': 25, 'obliques': 25, 'quads': 15, 'calves': 25,
            'trapezius': 20, 'back': 15, 'lats': 15, 'glutes': 15, 'hamstrings': 15,
            'forearms': 25, 'rear delts': 20
        }
        self.fatigue_threshold = 100  # Maximum fatigue level
        self.fatigue_multiplier = 0.1  # Reduce fatigue increase
        self.muscle_recovery = self.initialize_muscle_recovery()

        self.engagement_map = {
            'bench press': {'chest': 0.8, 'triceps': 0.2, 'shoulders': 0.2},
            'incline db press': {'chest': 0.7, 'shoulders': 0.3, 'triceps': 0.2},
            'db bench press': {'chest': 0.8, 'triceps': 0.2, 'shoulders': 0.2},
            'decline bench press': {'chest': 0.8, 'triceps': 0.2, 'shoulders': 0.2},
            'pec fly machine': {'chest': 1.0},
            'cable chest fly': {'chest': 1.0},
            'dip': {'chest': 0.6, 'triceps': 0.4},
            'shoulder press': {'shoulders': 0.8, 'triceps': 0.2},
            'db lateral raise': {'shoulders': 1.0},
            'face pull': {'shoulders': 0.6, 'trapezius': 0.4, 'rear delts': 0.3},
            'overhead press': {'shoulders': 0.8, 'triceps': 0.2},
            'front raise': {'shoulders': 1.0},
            'lat raises': {'shoulders': 1.0},
            'rear delt fly': {'shoulders': 1.0},
            'bicep curl': {'biceps': 1.0},
            'hammer curl': {'biceps': 0.8, 'forearms': 0.2},
            'preacher curl': {'biceps': 1.0},
            'cable curl': {'biceps': 1.0},
            'ez bar curl': {'biceps': 0.9, 'forearms': 0.1},
            'skull crushers': {'triceps': 1.0},
            'tricep extension': {'triceps': 1.0},
            'triceps pushdown': {'triceps': 1.0},
            'triceps rope pulldown': {'triceps': 1.0},
            'squat': {'quads': 0.6, 'glutes': 0.3, 'hamstrings': 0.2},
            'back squat': {'quads': 0.6, 'glutes': 0.3, 'hamstrings': 0.2},
            'front squat': {'quads': 0.7, 'glutes': 0.2, 'hamstrings': 0.1},
            'leg press': {'quads': 0.7, 'glutes': 0.2, 'hamstrings': 0.1},
            'hack squat': {'quads': 0.8, 'glutes': 0.2},
            'bulgarian split squat': {'quads': 0.5, 'glutes': 0.3, 'hamstrings': 0.2},
            'leg extension': {'quads': 1.0},
            'leg curl': {'hamstrings': 1.0},
            'romanian deadlift': {'hamstrings': 0.6, 'glutes': 0.3, 'back': 0.2},
            'prone machine hamstring curl': {'hamstrings': 1.0},
            'calf raise': {'calves': 1.0},
            'deadlift': {'back': 0.5, 'glutes': 0.3, 'hamstrings': 0.3, 'trapezius': 0.2, 'forearms': 0.1},
            'deadlift paused': {'back': 0.5, 'glutes': 0.3, 'hamstrings': 0.3, 'trapezius': 0.2, 'forearms': 0.1},
            'lat pulldown': {'lats': 0.8, 'biceps': 0.2},
            'pull up': {'lats': 0.7, 'biceps': 0.3},
            'row': {'back': 0.7, 'biceps': 0.3},
            'barbell row': {'back': 0.7, 'biceps': 0.2, 'forearms': 0.1},
            't bar row': {'back': 0.8, 'biceps': 0.2},
            'seated cable row': {'back': 0.7, 'biceps': 0.2, 'forearms': 0.1},
            'back extension': {'back': 0.8, 'glutes': 0.2},
            'crunch': {'abs': 1.0},
            'plank': {'abs': 0.7, 'obliques': 0.3},
            'reverse fly': {'rear delts': 0.8, 'trapezius': 0.2},
            'bent over lateral raise': {'rear delts': 0.9, 'trapezius': 0.1},
        }

        self.prs = self.initialize_prs()

        self.date_entry = None
        self.calculate_weekly_muscle_engagement()
        self.create_ui()
        self.update_recovery()
        self.recalculate_prs()

    def initialize_prs(self):
        """Initialize personal records (PRs) for exercises."""
        prs = {}
        for week in self.workouts:
            for day in week:
                for exercise in day['exercises']:
                    exercise_name = exercise['exercise']
                    if exercise_name not in prs:
                        prs[exercise_name] = {
                            'weight': {'value': 0, 'date': None},
                            'reps': {'value': 0, 'date': None},
                            'history': []
                        }
        return prs

    def update_prs(self, exercise):
        """Update PRs based on the current exercise data."""
        exercise_name = exercise['exercise']
        try:
            weight = float(exercise['weight'])
            reps = int(exercise['reps'])
        except ValueError:
            return  # Skip if weight or reps are not valid numbers

        if exercise_name not in self.prs:
            self.prs[exercise_name] = {
                'weight': {'value': 0, 'date': None},
                'reps': {'value': 0, 'date': None},
                'history': []
            }

        pr_data = self.prs[exercise_name]

        # Update weight PR
        if weight > pr_data['weight']['value']:
            pr_data['weight'] = {'value': weight, 'date': self.current_date}

        # Update reps PR for the current weight
        if weight == pr_data['weight']['value'] and reps > pr_data['reps']['value']:
            pr_data['reps'] = {'value': reps, 'date': self.current_date}

        # Add to history
        pr_data['history'].append({'date': self.current_date, 'weight': weight, 'reps': reps})

    def show_pr_details(self, exercise_name):
        """
        Display PR details for the given exercise.
        """
        exercise_name_lower = exercise_name.lower()

        # Specific matching logic for different exercises
        if 'squat' in exercise_name_lower:
            matching_exercises = [
                key for key in self.prs.keys()
                if 'squat' in key.lower() and 'hack' not in key.lower() and 'bulgarian split' not in key.lower()
            ]
        elif 'bench' in exercise_name_lower:
            matching_exercises = [
                key for key in self.prs.keys()
                if 'bench' in key.lower()
            ]
        elif 'deadlift' in exercise_name_lower:
            matching_exercises = [
                key for key in self.prs.keys()
                if 'deadlift' in key.lower()
            ]
        else:
            # Exact match fallback
            matching_exercises = [
                key for key in self.prs.keys()
                if exercise_name_lower == key.lower()
            ]

        if not matching_exercises:
            messagebox.showinfo("PR Details", f"No PR data available for {exercise_name}")
            return

        # Group by base exercise name
        condensed_prs = {}
        for matching_exercise in matching_exercises:
            if "bench" in matching_exercise.lower() or "barbell" in matching_exercise.lower():
                if "incline" in matching_exercise.lower():
                    base_name = "incline bench press"
                elif "decline" in matching_exercise.lower():
                    base_name = "decline bench press"
                elif "db" in matching_exercise.lower():
                    base_name = "db bench press"
                else:
                    base_name = "bench press"
            else:
                base_name = matching_exercise.split()[0].lower()

            pr_data = self.prs[matching_exercise]

            if pr_data['weight']['value'] > 0:  # Ignore entries with 0 PRs
                if base_name not in condensed_prs:
                    condensed_prs[base_name] = pr_data
                else:
                    # Compare and keep the highest PR
                    if pr_data['weight']['value'] > condensed_prs[base_name]['weight']['value']:
                        condensed_prs[base_name] = pr_data
                    elif pr_data['weight']['value'] == condensed_prs[base_name]['weight']['value'] and pr_data['reps']['value'] > condensed_prs[base_name]['reps']['value']:
                        condensed_prs[base_name] = pr_data

        details = f"PRs for variations of '{exercise_name}':\n\n"
        for base_name, pr_data in condensed_prs.items():
            # Capitalize the base name for display
            display_name = " ".join([word.capitalize() for word in base_name.split()])

            details += f"{display_name}:\n"
            details += f" - Weight PR: {pr_data['weight']['value']} kg on {pr_data['weight']['date']}\n"
            details += f" - Reps PR: {pr_data['reps']['value']} reps on {pr_data['reps']['date']}\n"
            details += "\n"

        messagebox.showinfo("PR Details", details)

    def show_progress_graph(self, exercise_group, exercise_names=None):
        """
        Show a progress graph for the given exercise group.
        """
        if exercise_names is None:
            exercise_names = [exercise_group]

        dates = []
        weights = []

        for exercise_name in exercise_names:
            for pr_exercise_name in self.prs.keys():
                if exercise_name.lower() == pr_exercise_name.lower() and self.prs[pr_exercise_name]['history']:
                    history = self.prs[pr_exercise_name]['history']

                    for entry in history:
                        if isinstance(entry['date'], str):
                            dates.append(datetime.strptime(entry['date'], '%Y-%m-%d').date())
                        elif isinstance(entry['date'], date):
                            dates.append(entry['date'])
                        else:
                            continue  # Skip if date is in an unexpected format
                        weights.append(entry['weight'])

        if not dates or not weights:
            messagebox.showinfo("Progress Graph", f"No progress data available for {exercise_group}.")
            return

        # Sort the dates and weights based on the dates
        sorted_data = sorted(zip(dates, weights))
        sorted_dates, sorted_weights = zip(*sorted_data)

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.plot(sorted_dates, sorted_weights, marker='o')
        ax.set_title(f"{exercise_group} Progress")
        ax.set_xlabel("Date")
        ax.set_ylabel("Weight (kg)")
        ax.grid(True)

        # Rotate and align the tick labels so they look better
        fig.autofmt_xdate()

        # Use a tkinter window to display the matplotlib figure
        window = tk.Toplevel(self.root)
        window.title(f"{exercise_group} Progress")
        canvas = FigureCanvasTkAgg(fig, master=window)
        canvas.draw()
        canvas.get_tk_widget().pack()

    def recalculate_prs(self):
        """Recalculate PRs based on the current workouts."""
        self.prs = self.initialize_prs()
        for week in self.workouts:
            for day in week:
                self.current_date = day.get('date', self.current_date)
                for exercise in day['exercises']:
                    self.update_prs(exercise)

    def initialize_weekly_muscle_engagement(self):
        """Initialize weekly muscle engagement data."""
        return {week: self.initialize_muscle_engagement() for week in range(1, self.total_weeks + 1)}

    def calculate_weekly_muscle_engagement(self):
        """Calculate weekly muscle engagement for all weeks."""
        for week in range(1, self.total_weeks + 1):
            self.weekly_muscle_engagement[week] = self.initialize_muscle_engagement()
            for day, workout in enumerate(self.workouts[week - 1], start=1):
                workout_date = datetime.strptime(workout.get('date', '2000-01-01'), '%Y-%m-%d').date()
                for exercise in workout['exercises']:
                    self.engage_muscles(exercise, workout_date)

    def initialize_muscle_recovery(self):
        """Initialize muscle recovery data."""
        return {muscle: {'fatigue': 0, 'last_workout': None} for muscle in self.recovery_rates}

    def initialize_muscle_engagement(self):
        return {
            'chest': 0, 'shoulders': 0, 'biceps': 0, 'triceps': 0,
            'abs': 0, 'obliques': 0, 'quads': 0, 'calves': 0,
            'trapezius': 0, 'back': 0, 'lats': 0, 'glutes': 0, 'hamstrings': 0,
            'forearms': 0, 'rear delts': 0
        }

    def find_start_date(self):
        """Find the earliest date in the workouts."""
        start_date = None
        for week in self.workouts:
            for day in week:
                if 'date' in day and day['date']:
                    date_obj = datetime.strptime(day['date'], '%Y-%m-%d').date()
                    if start_date is None or date_obj < start_date:
                        start_date = date_obj
        return start_date or date.today()

    def calculate_current_fatigue(self):
        """Calculate current fatigue levels for all muscles."""
        for muscle, data in self.muscle_recovery.items():
            last_workout = data['last_workout']
            if last_workout:
                days_since_workout = max(0, (self.current_date - last_workout).days)
                recovery_rate = self.recovery_rates[muscle]
                recovery_amount = recovery_rate * days_since_workout
                current_fatigue = max(0, data['fatigue'] - recovery_amount)
                self.muscle_recovery[muscle]['fatigue'] = current_fatigue

    def create_ui(self):
        """Create the user interface."""
        # Main frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Top frame
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=(0, 10))

        # Week selection
        ttk.Label(top_frame, text="Select Week:").pack(side=tk.LEFT)
        self.week_var = tk.StringVar()
        self.week_dropdown = ttk.Combobox(top_frame, textvariable=self.week_var, state="readonly")
        self.week_dropdown.pack(side=tk.LEFT, padx=(0, 10))
        self.week_dropdown.bind("<<ComboboxSelected>>", self.on_week_select)

        # Date entry
        self.date_entry = DateEntry(top_frame, date_pattern='yyyy-mm-dd', width=10)
        self.date_entry.pack(side=tk.LEFT)

        # Control buttons
        ttk.Button(top_frame, text="Save Date", command=self.save_current_date).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(top_frame, text="Previous Day", command=self.previous_day).pack(side=tk.LEFT)
        ttk.Button(top_frame, text="Next Day", command=self.next_day).pack(side=tk.LEFT)
        ttk.Button(top_frame, text="Swap Days", command=self.swap_days).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(top_frame, text="Week Overview", command=self.show_week_overview).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(top_frame, text="Add Day", command=self.add_day).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(top_frame, text="Remove Day", command=self.remove_day).pack(side=tk.LEFT, padx=(0, 10))

        # Labels
        self.week_label = ttk.Label(top_frame, text="Week 1", font=('Arial', 12, 'bold'))
        self.week_label.pack(side=tk.LEFT, padx=20)

        self.day_label = ttk.Label(top_frame, text="Day 1", font=('Arial', 12, 'bold'))
        self.day_label.pack(side=tk.LEFT)

        self.rpe_label = ttk.Label(top_frame, text="RPE: 60.0%", font=('Arial', 12))
        self.rpe_label.pack(side=tk.RIGHT)

        # Content frame
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)

        # Workout frame
        self.workout_frame = ttk.Frame(content_frame)
        self.workout_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Treeview for exercises
        self.tree = ttk.Treeview(self.workout_frame, columns=(
            'Exercise', 'Sets', 'Reps', 'Weight', 'Load', 'Weight PR', 'Reps PR'
        ), show='headings', height=15)
        self.tree.heading('Exercise', text='Exercise')
        self.tree.heading('Sets', text='Sets')
        self.tree.heading('Reps', text='Reps')
        self.tree.heading('Weight', text='Weight')
        self.tree.heading('Load', text='Load')  
        self.tree.heading('Weight PR', text='Weight PR')
        self.tree.heading('Reps PR', text='Reps PR')

        self.tree.column('Exercise', width=200)
        self.tree.column('Sets', width=50, anchor='center')
        self.tree.column('Reps', width=50, anchor='center')
        self.tree.column('Weight', width=70, anchor='center')
        self.tree.column('Load', width=70, anchor='center')
        self.tree.column('Weight PR', width=70, anchor='center')
        self.tree.column('Reps PR', width=70, anchor='center')

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Scrollbar
        scrollbar = ttk.Scrollbar(self.workout_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Treeview bindings
        self.tree.bind('<Double-1>', self.on_double_click)
        self.tree.bind('<ButtonPress-1>', self.on_press)
        self.tree.bind('<B1-Motion>', self.on_motion)
        self.tree.bind('<ButtonRelease-1>', self.on_release)


        
        
        # Muscle visualization frame
        muscle_frame = ttk.Frame(content_frame, width=650)
        muscle_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))

        self.muscle_image = Image.open("muscle_diagram.png")
        self.muscle_image = self.muscle_image.resize((600, 400), Image.Resampling.LANCZOS)
        self.muscle_photo = ImageTk.PhotoImage(self.muscle_image)
        self.muscle_canvas = tk.Canvas(muscle_frame, width=600, height=400)
        self.muscle_canvas.pack(side=tk.TOP)
        self.muscle_canvas.create_image(300, 200, image=self.muscle_photo, tags="muscle_overlay")
        self.muscle_canvas.image = self.muscle_photo  # Prevent garbage collection

        # Buttons under the image
        muscle_buttons_frame = ttk.Frame(muscle_frame)
        muscle_buttons_frame.pack(side=tk.TOP, pady=(10, 0))

        buttons = [
            ("Add Exercise", self.add_exercise),
            ("Remove Exercise", self.remove_exercise),
            ("Save Data", self.save_data),
            ("Load Data", self.load_data),
            ("Week Overview", self.show_week_overview),
            ("Show Bench PR Details", lambda: self.show_pr_details("Bench Press")),
            ("Show Squat PR Details", lambda: self.show_pr_details("Squats")),
            ("Show Deadlift PR Details", lambda: self.show_pr_details("Deadlift")),
            ("Show Squat Progress Graph", lambda: self.show_progress_graph("Squats", ["squat", "back squat", "front squat"])),
            ("Show Bench Progress Graph", lambda: self.show_progress_graph("Bench Presses", ["bench press", "barbell bench press"])),
            ("Show Deadlift Progress Graph", lambda: self.show_progress_graph("Deadlift")),
        ]

        button_options = {'width': 30}

        for index, (text, command) in enumerate(buttons):
            row = index // 2
            column = index % 2
            ttk.Button(muscle_buttons_frame, text=text, command=command, **button_options).grid(row=row, column=column, padx=5, pady=5)

        muscle_buttons_frame.columnconfigure(0, weight=1)
        muscle_buttons_frame.columnconfigure(1, weight=1)

        # Bottom frame (keep if you have other widgets)
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))

        self.volume_load_label = ttk.Label(bottom_frame, text="Total Volume Load: 0", font=('Segoe UI', 12))
        self.volume_load_label.pack(side=tk.RIGHT)

        self.update_week_dropdown()
        self.load_workout_data()

    def add_day(self):
        """Add a new day to the current week."""
        max_days_per_week = 7
        current_week_days = self.workouts[self.current_week - 1]

        if len(current_week_days) >= max_days_per_week:
            messagebox.showwarning("Maximum Days Reached", "A week can have a maximum of 7 days.")
            return

        new_day_number = sum(len(week) for week in self.workouts[:self.current_week - 1]) + len(current_week_days) + 1
        new_day = {
            'day': f"Day {new_day_number}",
            'exercises': [],
            'date': (date.today() + timedelta(days=new_day_number - 1)).strftime('%Y-%m-%d')
        }
        current_week_days.append(new_day)
        self.current_day = len(current_week_days) - 1
        self.renumber_all_days()
        self.load_workout_data()
        self.update_week_dropdown()
        messagebox.showinfo("Day Added", f"Added {new_day['day']} to the current week and renumbered subsequent days.")

    def remove_day(self):
        """Remove a day from the current week."""
        current_week_days = self.workouts[self.current_week - 1]

        if len(current_week_days) <= 1:
            messagebox.showwarning("Cannot Remove Day", "A week must have at least one day.")
            return

        remove_dialog = tk.Toplevel(self.root)
        remove_dialog.title("Remove Day")

        ttk.Label(remove_dialog, text="Select day to remove:").pack(pady=5)

        day_var = tk.StringVar(remove_dialog)
        day_options = [day['day'] for day in current_week_days]
        day_dropdown = ttk.Combobox(remove_dialog, textvariable=day_var, values=day_options, state="readonly")
        day_dropdown.pack(pady=5)
        day_dropdown.set(day_options[0])

        def perform_remove():
            day_index = day_options.index(day_var.get())
            removed_day = current_week_days.pop(day_index)
            self.renumber_all_days()
            if self.current_day >= len(current_week_days):
                self.current_day = len(current_week_days) - 1
            self.load_workout_data()
            self.update_week_dropdown()
            remove_dialog.destroy()
            messagebox.showinfo("Day Removed", f"Removed {removed_day['day']} from the current week and renumbered subsequent days.")

        ttk.Button(remove_dialog, text="Remove", command=perform_remove).pack(pady=10)

    def renumber_all_days(self):
        """Renumber all days across all weeks."""
        day_number = 1
        for week in self.workouts:
            for workout in week:
                workout['day'] = f"Day {day_number}"
                day_number += 1

    def swap_days(self):
        """Swap two days within the current week."""
        swap_dialog = tk.Toplevel(self.root)
        swap_dialog.title("Swap Days")

        ttk.Label(swap_dialog, text="Select days to swap:").grid(row=0, column=0, columnspan=2, pady=5)

        day1_var = tk.StringVar(swap_dialog)
        day2_var = tk.StringVar(swap_dialog)

        day_options = [day['day'] for day in self.workouts[self.current_week - 1]]

        day1_dropdown = ttk.Combobox(swap_dialog, textvariable=day1_var, values=day_options, state="readonly")
        day1_dropdown.grid(row=1, column=0, padx=5, pady=5)
        day1_dropdown.set(day_options[0])

        day2_dropdown = ttk.Combobox(swap_dialog, textvariable=day2_var, values=day_options, state="readonly")
        day2_dropdown.grid(row=1, column=1, padx=5, pady=5)
        day2_dropdown.set(day_options[1])

        def perform_swap():
            day1_index = day_options.index(day1_var.get())
            day2_index = day_options.index(day2_var.get())

            if day1_index != day2_index:
                # Swap the workout data
                week = self.workouts[self.current_week - 1]
                week[day1_index], week[day2_index] = week[day2_index], week[day1_index]
                # Renumber days
                self.renumber_all_days()
                self.load_workout_data()
                swap_dialog.destroy()
                messagebox.showinfo("Swap Successful", "The selected days have been swapped and day numbers updated.")

        ttk.Button(swap_dialog, text="Swap", command=perform_swap).grid(row=2, column=0, columnspan=2, pady=10)

    def show_week_overview(self):
        """Display an overview of the current week's workouts."""
        overview_window = tk.Toplevel(self.root)
        overview_window.title(f"Week {self.current_week} Overview")
        overview_window.geometry("1000x600")

        main_frame = ttk.Frame(overview_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        for day_index, workout in enumerate(self.workouts[self.current_week - 1]):
            day_frame = ttk.LabelFrame(scrollable_frame, text=workout['day'])
            day_frame.pack(fill=tk.X, expand=True, padx=5, pady=5)

            tree = ttk.Treeview(day_frame, columns=('Exercise', 'Sets', 'Reps', 'Weight'), show='headings', height=len(workout['exercises']))
            tree.heading('Exercise', text='Exercise')
            tree.heading('Sets', text='Sets')
            tree.heading('Reps', text='Reps')
            tree.heading('Weight', text='Weight')

            tree.column('Exercise', width=200)
            tree.column('Sets', width=50, anchor='center')
            tree.column('Reps', width=50, anchor='center')
            tree.column('Weight', width=70, anchor='center')

            for exercise in workout['exercises']:
                tree.insert('', 'end', values=(exercise['exercise'], exercise['sets'], exercise['reps'], exercise['weight']))

            tree.pack(fill=tk.X, expand=True)

            ttk.Button(day_frame, text="Edit", command=lambda idx=day_index: self.edit_day(idx)).pack(side=tk.LEFT, padx=5, pady=5)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def edit_day(self, day_index):
        """Edit a specific day by setting it as the current day."""
        self.current_day = day_index
        self.load_workout_data()

    def save_current_date(self):
        """Save the selected date for the current workout day."""
        if self.current_week <= len(self.workouts) and self.current_day < len(self.workouts[self.current_week - 1]):
            selected_date = self.date_entry.get_date().strftime('%Y-%m-%d')
            self.workouts[self.current_week - 1][self.current_day]['date'] = selected_date

    def update_week_dropdown(self):
        """Update the week selection dropdown."""
        weeks = [f"Week {i+1}" for i in range(len(self.workouts))]
        self.week_dropdown['values'] = weeks
        if self.current_week <= len(weeks):
            self.week_dropdown.set(f"Week {self.current_week}")
        else:
            self.week_dropdown.set(weeks[-1] if weeks else "")

    def on_week_select(self, event):
        """Handle week selection change."""
        selected_week = int(self.week_var.get().split()[1])
        if selected_week != self.current_week:
            self.current_week = selected_week
            self.current_day = 0
            self.load_workout_data()

    def load_workout_data(self):
        """Load workout data for the current week and day."""
        if not self.workouts:
            messagebox.showwarning("No Data", "No workout data available.")
            return

        self.total_weeks = len(self.workouts)
        self.update_week_dropdown()

        if self.current_week <= len(self.workouts) and self.current_day < len(self.workouts[self.current_week - 1]):
            workout = self.workouts[self.current_week - 1][self.current_day]
            workout_date = workout.get('date', None)
            if workout_date:
                self.current_date = datetime.strptime(workout_date, '%Y-%m-%d').date()
            else:
                self.current_date = date.today()
            self.date_entry.set_date(self.current_date)

            self.tree.delete(*self.tree.get_children())
            self.muscle_recovery = self.initialize_muscle_recovery()

            # Calculate fatigue up to the current date
            for week in self.workouts:
                for day_workout in week:
                    day_date = datetime.strptime(day_workout.get('date', '2000-01-01'), '%Y-%m-%d').date()
                    if day_date <= self.current_date:
                        for exercise in day_workout['exercises']:
                            self.engage_muscles(exercise, day_date)

            self.calculate_current_fatigue()

            for exercise in workout['exercises']:
                self.update_prs(exercise)
                load = self.calculate_load(exercise)
                weight_pr = self.prs[exercise['exercise']]['weight']['value']
                reps_pr = self.prs[exercise['exercise']]['reps']['value']
                self.tree.insert('', 'end', values=(
                    exercise['exercise'],
                    exercise['sets'],
                    exercise['reps'],
                    exercise['weight'],
                    f"{load:.0f}",
                    f"{weight_pr:.1f}",
                    f"{reps_pr}"
                ))

            self.week_label.config(text=f"Week {self.current_week}")
            self.day_label.config(text=workout['day'])
            self.rpe_label.config(text=f"RPE: {self.calculate_rpe(self.current_week):.1f}%")
            self.update_muscle_recovery()
            self.update_volume_load()
            self.update_muscle_visualization()
        else:
            messagebox.showwarning("Invalid Week/Day", "The selected week or day is out of range.")

    def next_day(self):
        """Navigate to the next workout day."""
        self.save_current_date()
        if self.current_day < len(self.workouts[self.current_week - 1]) - 1:
            self.current_day += 1
        elif self.current_week < len(self.workouts):
            self.current_week += 1
            self.current_day = 0
        self.load_workout_data()

    def previous_day(self):
        """Navigate to the previous workout day."""
        self.save_current_date()
        if self.current_day > 0:
            self.current_day -= 1
        elif self.current_week > 1:
            self.current_week -= 1
            self.current_day = len(self.workouts[self.current_week - 1]) - 1
        self.load_workout_data()

    def add_exercise(self):
        """Add a new exercise to the current day."""
        if self.current_week <= len(self.workouts) and self.current_day < len(self.workouts[self.current_week - 1]):
            new_exercise = {'exercise': 'New Exercise', 'sets': '0', 'reps': '0', 'weight': '0'}
            self.workouts[self.current_week - 1][self.current_day]['exercises'].append(new_exercise)
            self.tree.insert('', 'end', values=(new_exercise['exercise'], new_exercise['sets'], new_exercise['reps'], new_exercise['weight'], '0'))
            self.update_volume_load()
        else:
            messagebox.showwarning("Invalid Week/Day", "Cannot add exercise to the selected week/day.")

    def remove_exercise(self):
        """Remove the selected exercise from the current day."""
        selected_item = self.tree.selection()
        if selected_item and self.current_week <= len(self.workouts) and self.current_day < len(self.workouts[self.current_week - 1]):
            index = self.tree.index(selected_item)
            self.tree.delete(selected_item)
            del self.workouts[self.current_week - 1][self.current_day]['exercises'][index]
            self.update_volume_load()

    def calculate_load(self, exercise):
        """Calculate the load for a given exercise."""
        try:
            sets = int(exercise['sets'])
            reps = int(exercise['reps'])
            weight = float(exercise['weight'])
            return sets * reps * weight
        except ValueError:
            return 0

    def update_volume_load(self):
        """Update the total volume load."""
        total_load = sum(
            self.calculate_load(exercise)
            for week in self.workouts
            for day in week
            for exercise in day['exercises']
        )
        self.volume_load_label.config(text=f"Total Volume Load: {total_load:.0f}")
        self.calculate_muscle_engagement()

    def save_data(self):
        """Save workout data and PRs to a JSON file."""
        file_path = filedialog.asksaveasfilename(defaultextension=".json")
        if file_path:
            try:
                data_to_save = {
                    'workouts': self.workouts,
                    'prs': self.prs
                }
                data_to_save = self.prepare_data_for_json(data_to_save)
                with open(file_path, 'w') as file:
                    json.dump(data_to_save, file, indent=2)
                messagebox.showinfo("Save Successful", "Workout data and PRs have been saved successfully.")
                self.update_week_dropdown()
            except Exception as e:
                messagebox.showerror("Save Error", f"An error occurred while saving: {str(e)}")

    @staticmethod
    def prepare_data_for_json(data):
        """Prepare data for JSON serialization."""
        if isinstance(data, dict):
            return {k: WorkoutPlanner.prepare_data_for_json(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [WorkoutPlanner.prepare_data_for_json(item) for item in data]
        elif isinstance(data, (date, datetime)):
            return data.isoformat()
        else:
            return data

    def load_data(self):
        """Load workout data and PRs from a JSON file."""
        file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if file_path:
            with open(file_path, 'r') as file:
                loaded_data = json.load(file)

            if isinstance(loaded_data, dict) and 'workouts' in loaded_data:
                self.workouts = self.restructure_workouts(loaded_data['workouts'])
                self.prs = loaded_data.get('prs', self.initialize_prs())
            else:
                self.workouts = self.restructure_workouts(loaded_data)
                self.prs = self.initialize_prs()

            self.current_week = 1
            self.current_day = 0
            self.recalculate_prs()
            self.load_workout_data()
            self.update_week_dropdown()
            messagebox.showinfo("Load Successful", "Workout data has been loaded successfully.")

    def export_to_excel(self):
        """Export the current week's workouts to an Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Week {self.current_week}"

        header_font = Font(bold=True)
        centered = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Day', 'Exercise', 'Sets', 'Reps', 'Weight']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = centered
            cell.border = border

        row = 2
        for day, workout in enumerate(self.workouts[self.current_week - 1], start=1):
            exercises = workout['exercises']
            ws.merge_cells(start_row=row, start_column=1, end_row=row + len(exercises) - 1, end_column=1)
            ws.cell(row=row, column=1, value=workout['day']).alignment = centered

            for exercise in exercises:
                ws.cell(row=row, column=2, value=exercise['exercise'])
                ws.cell(row=row, column=3, value=exercise['sets']).alignment = centered
                ws.cell(row=row, column=4, value=exercise['reps']).alignment = centered
                ws.cell(row=row, column=5, value=exercise['weight']).alignment = centered
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if file_path:
            wb.save(file_path)

    def engage_muscles(self, exercise, workout_date):
        """Engage muscles based on the exercise performed."""
        exercise_name = exercise['exercise'].lower()
        load = self.calculate_load(exercise)

        matched_exercise = next((key for key in self.engagement_map.keys() if key in exercise_name), None)

        if matched_exercise and load > 0:
            engagement = self.engagement_map[matched_exercise]
            for muscle, factor in engagement.items():
                if muscle not in self.muscle_recovery:
                    self.muscle_recovery[muscle] = {'fatigue': 0, 'last_workout': workout_date}

                fatigue_increase = min(load * factor * self.fatigue_multiplier / 100, self.fatigue_threshold)
                current_fatigue = self.muscle_recovery[muscle]['fatigue']
                new_fatigue = min(self.fatigue_threshold, current_fatigue + fatigue_increase)

                self.muscle_recovery[muscle]['fatigue'] = new_fatigue
                self.muscle_recovery[muscle]['last_workout'] = workout_date

    def calculate_muscle_engagement(self):
        """Calculate muscle engagement for the current week."""
        self.weekly_muscle_engagement = {}
        for week in range(1, self.total_weeks + 1):
            self.weekly_muscle_engagement[week] = self.initialize_muscle_engagement()
            for day, workout_day in enumerate(self.workouts[week - 1], start=1):
                workout_date = datetime.strptime(workout_day.get('date', '2000-01-01'), '%Y-%m-%d').date()
                if workout_date <= self.current_date:
                    for exercise in workout_day['exercises']:
                        self.engage_muscles(exercise, workout_date)
        self.update_muscle_visualization()

    def on_focus_out(self, event, tree, item, column):
        """Handle focus out event when editing a cell."""
        new_value = event.widget.get()
        tree.set(item, column, new_value)
        event.widget.destroy()

        if self.current_week <= len(self.workouts) and self.current_day < len(self.workouts[self.current_week - 1]):
            exercise_index = tree.index(item)
            column_name = tree.heading(column)['text'].lower()
            self.workouts[self.current_week - 1][self.current_day]['exercises'][exercise_index][column_name] = new_value

            # Recalculate load and update PRs
            exercise = self.workouts[self.current_week - 1][self.current_day]['exercises'][exercise_index]
            load = self.calculate_load(exercise)
            tree.set(item, 'Load', f"{load:.0f}")

            self.update_prs(exercise)
            weight_pr = self.prs[exercise['exercise']]['weight']['value']
            reps_pr = self.prs[exercise['exercise']]['reps']['value']
            tree.set(item, 'Weight PR', f"{weight_pr:.1f}")
            tree.set(item, 'Reps PR', f"{reps_pr}")

            self.update_volume_load()

    def update_muscle_recovery(self):
        """Update muscle recovery data based on fatigue and recovery rates."""
        for muscle, recovery_data in self.muscle_recovery.items():
            last_workout_date = recovery_data['last_workout']
            if last_workout_date:
                days_elapsed = (self.current_date - last_workout_date).days
                recovery_rate = self.recovery_rates[muscle]
                recovery_amount = recovery_rate * days_elapsed

                current_fatigue = recovery_data['fatigue']
                new_fatigue = max(current_fatigue - recovery_amount, 0.0)

                self.muscle_recovery[muscle]['fatigue'] = new_fatigue

    def update_recovery(self):
        """Periodic update of muscle recovery."""
        self.update_muscle_recovery()
        self.update_muscle_visualization()
        self.root.after(86400000, self.update_recovery)  # Update every 24 hours (in milliseconds)

    def on_press(self, event):
        """Handle mouse press event on the treeview for drag-and-drop."""
        tree = event.widget
        item = tree.identify_row(event.y)
        if item:
            tree.selection_set(item)
            self.drag_data = {'item': item, 'index': tree.index(item)}

    def on_motion(self, event):
        """Handle mouse motion event for drag-and-drop."""
        tree = event.widget
        if hasattr(self, 'drag_data'):
            item = tree.identify_row(event.y)
            if item and item != self.drag_data['item']:
                tree.move(self.drag_data['item'], tree.parent(item), tree.index(item))

    def on_release(self, event):
        """Handle mouse release event to finalize drag-and-drop."""
        tree = event.widget
        if hasattr(self, 'drag_data') and self.current_week <= len(self.workouts) and self.current_day < len(self.workouts[self.current_week - 1]):
            new_index = tree.index(self.drag_data['item'])
            exercise = self.workouts[self.current_week - 1][self.current_day]['exercises'].pop(self.drag_data['index'])
            self.workouts[self.current_week - 1][self.current_day]['exercises'].insert(new_index, exercise)
            del self.drag_data

    def on_double_click(self, event):
        """Handle double-click event to edit a cell."""
        tree = event.widget
        column = tree.identify_column(event.x)
        item = tree.identify_row(event.y)

        if column in ('#1', '#2', '#3', '#4'):  # Editable columns
            tree.item(item, open=True)
            x, y, width, height = tree.bbox(item, column)

            entry_edit = ttk.Entry(tree, width=width // 8)
            entry_edit.place(x=x, y=y, width=width, height=height)
            entry_edit.insert(0, tree.set(item, column))
            entry_edit.select_range(0, tk.END)
            entry_edit.focus()
            entry_edit.bind('<FocusOut>', lambda e, t=tree, i=item, c=column: self.on_focus_out(e, t, i, c))
            entry_edit.bind('<Return>', lambda e, t=tree, i=item, c=column: self.on_focus_out(e, t, i, c))

    def update_muscle_visualization(self):
        """Update the muscle visualization based on current fatigue levels."""
        heat_map = Image.new('RGBA', (600, 400), (255, 255, 255, 0))
        draw = ImageDraw.Draw(heat_map)

        muscles = {
            'chest': [(145, 85, 220, 135)],
            'shoulders': [(130, 85, 145, 110), (220, 85, 235, 110)],
            'biceps': [(120, 110, 150, 140), (216, 110, 245, 140)],
            'abs': [(165, 115, 200, 180)],
            'obliques': [(150, 115, 165, 180), (200, 115, 215, 180)],
            'quads': [(140, 170, 180, 250), (187, 170, 230, 250)],
            'calves': [(140, 270, 175, 290), (195, 270, 230, 290)],
            'back': [(385, 75, 465, 160)],
            'triceps': [(360, 100, 385, 140), (465, 100, 490, 140)],
            'lats': [(385, 110, 420, 160), (430, 110, 465, 160)],
            'trapezius': [(405, 60, 445, 75)],
            'rear delts': [(370, 80, 385, 100), (465, 80, 480, 100)],
            'glutes': [(390, 160, 460, 210)],
            'hamstrings': [(380, 210, 420, 270), (430, 210, 470, 270)],
        }

        self.calculate_current_fatigue()

        for muscle, areas in muscles.items():
            if muscle in self.muscle_recovery:
                fatigue = self.muscle_recovery[muscle]['fatigue']
                color = self.get_recovery_color(fatigue)
                for area in areas:
                    draw.rectangle(area, fill=color)

        heat_map_photo = ImageTk.PhotoImage(heat_map)
        self.muscle_canvas.delete("heat_map")
        self.muscle_canvas.create_image(300, 200, image=heat_map_photo, tags="heat_map")
        self.muscle_canvas.tag_lower("heat_map", "muscle_overlay")
        self.heat_map_photo = heat_map_photo  # Prevent garbage collection

    def get_recovery_color(self, fatigue):
        """Get the color representing the recovery state based on fatigue."""
        fatigue = max(0, min(fatigue, self.fatigue_threshold))
        recovery = 1 - (fatigue / self.fatigue_threshold)
        r = int(255 * (1 - recovery))
        g = int(255 * recovery)
        b = 0
        return (r, g, b, 200)

    @staticmethod
    def calculate_rpe(week):
        """Calculate RPE based on the week number."""
        if week <= 6:
            return 60 + (week - 1) * 4.5
        elif week == 7:
            return 70
        else:
            return 75 + (week - 8) * 4.5

    def restructure_workouts(self, workouts):
        if isinstance(workouts, list) and all(isinstance(week, list) for week in workouts):
            return workouts
        
        restructured = []
        week = []
        for workout in workouts:
            week.append(workout)
            if len(week) == 5:  # Assuming 5 workouts per week
                restructured.append(week)
                week = []
        if week:  # Add any remaining workouts
            restructured.append(week)
        return restructured


if __name__ == "__main__":
    root = ThemedTk(theme="arc")  # Choose your desired theme
    root.title("Workout Plan")
    root.geometry("1400x900")

    try:
        with open('workouts.json', 'r') as file:
            workouts = json.load(file)
    except FileNotFoundError:
        with open('paste.txt', 'r') as file:
            workout_data = file.read()
        workouts = parse_workout_data(workout_data)

    app = WorkoutPlanner(root, workouts)
    root.mainloop()
